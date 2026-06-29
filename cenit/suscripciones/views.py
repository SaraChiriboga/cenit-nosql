import json
from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from datetime import datetime, timedelta
from bson import ObjectId
from django.conf import settings
from django.core.mail import EmailMessage
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML
from io import BytesIO

from cenit.mongo_client import db
from .models import MongoDoc, prepare_doc, parse_date

# ── Colecciones de MongoDB ──────────────────────────────
tipo_suscripciones_col = db['tipoSuscripciones']
promociones_col        = db['promociones']
suscripciones_col      = db['suscripciones']
notificaciones_col     = db['notificaciones']
playlists_col          = db['playlists']
estadisticas_col       = db['estadisticasDiarias']
usuarios_col           = db['usuarios']
canciones_col          = db['Cancion']


# ── Helpers ─────────────────────────────────────────────

def _get_or_404(collection, filtro, request, redirect_url, msg='No encontrado.'):
    """Busca un documento; si no existe, redirige con mensaje de error."""
    doc = collection.find_one(filtro)
    if not doc:
        messages.error(request, msg)
        return None
    return doc


def _usuarios_dropdown():
    """Devuelve lista de usuarios para los dropdowns de formularios."""
    docs = usuarios_col.find({}, {'id': 1, 'nombre': 1, 'apellido': 1}).sort('nombre', 1)
    return [MongoDoc(d) for d in docs]


def _planes_dropdown():
    """Devuelve lista de planes para los dropdowns."""
    return [MongoDoc(d) for d in tipo_suscripciones_col.find().sort('tipo_id', 1)]


def _promociones_dropdown(solo_activas=False):
    """Devuelve lista de promociones para los dropdowns."""
    filtro = {'estadoActivo': True} if solo_activas else {}
    return [MongoDoc(d) for d in promociones_col.find(filtro).sort('promo_id', 1)]


def _canciones_dropdown():
    """Devuelve lista de canciones para los dropdowns."""
    docs = canciones_col.find({}, {'cancion_id': 1, 'tituloCancion': 1}).sort('tituloCancion', 1)
    return [MongoDoc(d) for d in docs]


def _lookup_usuarios(user_ids):
    """Dado un set de IDs de usuario, retorna un dict {id: 'Nombre Apellido'}."""
    if not user_ids:
        return {}
    docs = usuarios_col.find({'id': {'$in': list(user_ids)}}, {'id': 1, 'nombre': 1, 'apellido': 1})
    return {d['id']: f"{d.get('nombre', '')} {d.get('apellido', '')}" for d in docs}


def _next_id(collection, field):
    """Genera el siguiente ID numérico auto-incremental para una colección."""
    last = collection.find_one(sort=[(field, -1)])
    return (last[field] + 1) if last else 1


def _parse_date_params(request):
    """Extrae filtros de fecha ?desde= y ?hasta= del GET, retorna (desde, hasta) datetime o None."""
    desde_str = request.GET.get('desde', '').strip()
    hasta_str = request.GET.get('hasta', '').strip()
    desde = None
    hasta = None
    try:
        if desde_str:
            desde = datetime.strptime(desde_str, '%Y-%m-%d')
    except (ValueError, TypeError):
        pass
    try:
        if hasta_str:
            hasta = datetime.strptime(hasta_str, '%Y-%m-%d') + timedelta(days=1)
    except (ValueError, TypeError):
        pass
    return desde, hasta, desde_str, hasta_str


def _paginate(request, docs):
    """Aplica ?page= y ?per_page= a una lista de docs. Retorna (slice, total, page, total_pages)."""
    page = request.GET.get('page', 1)
    try:
        page = int(page)
    except (ValueError, TypeError):
        page = 1
    per_page = request.GET.get('per_page', 50)
    try:
        per_page = int(per_page)
    except (ValueError, TypeError):
        per_page = 50
    if per_page > 200:
        per_page = 200
    total = len(docs)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    start = (page - 1) * per_page
    end = start + per_page
    return docs[start:end], total, page, total_pages, per_page


def _build_excel_response(wb, filename):
    """Toma un openpyxl Workbook y retorna un HttpResponse de descarga."""
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _estilo_excel_sheet(ws, headers):
    """Aplica formato de tabla nativa de Excel y estilos básicos de fuente/anchos."""
    from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment
    import re
    
    # 1. Definir rango de la tabla y añadir objeto Table nativo
    last_col = get_column_letter(len(headers))
    last_row = ws.max_row
    
    # Sanitizar displayName para evitar caracteres especiales inválidos en Excel
    clean_title = re.sub(r'[^a-zA-Z0-9_]', '', ws.title.replace(' ', '_'))
    if not clean_title:
        clean_title = "Tabla_Datos"
        
    displayName = f"Tabla_{clean_title}"
    table_range = f"A1:{last_col}{last_row}"
    
    # Crear y agregar la tabla
    tab = Table(displayName=displayName, ref=table_range)
    
    # IMPORTANTE: Inicializar las columnas explícitamente para que Excel active los filtros y ordenación
    tab.tableColumns = [TableColumn(id=i, name=h) for i, h in enumerate(headers, 1)]
    
    # TableStyleMedium9 tiene una gama cian/verde azulado excelente para Cénit
    style = TableStyleInfo(
        name="TableStyleMedium9", 
        showFirstColumn=False,
        showLastColumn=False, 
        showRowStripes=True, 
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    # 2. Aplicar fuentes, alineaciones y formatos básicos (Segoe UI y alturas)
    font_data = Font(name='Segoe UI', size=10)
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    
    # Alturas
    ws.row_dimensions[1].height = 26
    for row_idx in range(2, last_row + 1):
        ws.row_dimensions[row_idx].height = 20
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font_data
            
            # Alineación según contenido
            val = cell.value
            if isinstance(val, (int, float)) or (isinstance(val, str) and val.startswith('2026-')):
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
    # 3. Auto-ajustar anchos de columna
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)


# ── Helpers de datos para reportes (DRY) ────────────────

def _data_vencimientos(desde=None, hasta=None):
    """Retorna lista de suscripciones expiradas próximas a vencer."""
    filtro = {
        'estado': 'Expirada',
        'tipoSuscripcion.nombrePlan': {
            '$in': ['Premium Individual', 'Premium Duo', 'Premium Familiar', 'Premium Estudiante']
        }
    }
    if desde or hasta:
        filtro['fechaFin'] = {}
        if desde:
            filtro['fechaFin']['$gte'] = desde.isoformat()
        if hasta:
            filtro['fechaFin']['$lte'] = hasta.isoformat()
    docs = list(suscripciones_col.find(filtro).sort('fechaFin', 1))
    user_ids = {d.get('idUsuario') for d in docs if d.get('idUsuario')}
    user_names = _lookup_usuarios(user_ids)
    for d in docs:
        prepare_doc(d)
        resolved_name = user_names.get(d.get('idUsuario'), '—')
        d['usuario_nombre'] = resolved_name
        d['usuario_nombre_nombre'] = resolved_name
    return [MongoDoc(d) for d in docs]


def _data_promociones_vencidas(desde=None, hasta=None):
    """Retorna lista de promociones activas con alto descuento."""
    filtro = {'estadoActivo': True, 'porcentajeDesc': {'$gt': 20}}
    if desde or hasta:
        filtro['fechaExpira'] = {}
        if desde:
            filtro['fechaExpira']['$gte'] = desde.isoformat()
        if hasta:
            filtro['fechaExpira']['$lte'] = hasta.isoformat()
    docs = list(promociones_col.find(filtro).sort('fechaExpira', 1))
    for d in docs:
        prepare_doc(d)
    return [MongoDoc(d) for d in docs]


def _render_pdf(request, template_name, context, filename):
    """Genera un PDF desde un template y lo retorna como HttpResponse."""
    html_string = render_to_string(template_name, context, request=request)
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _dispatch_report_email(request, pdf_template, pdf_context, pdf_filename, excel_headers, write_excel_data_fn, excel_filename):
    """
    Genera y envía reportes por correo personalizando destinatario, mensaje y archivos adjuntos (PDF, Excel o ambos).
    """
    import json
    from django.core.mail import EmailMessage
    from weasyprint import HTML
    import openpyxl
    
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'error': 'Método no permitido'}, status=405)
        
    try:
        data = json.loads(request.body)
        destinatario = data.get('destinatario', request.user.email or 'admin@cenit.com')
        subject = data.get('asunto', 'Reporte Cénit')
        body = data.get('mensaje', 'Adjuntamos el informe solicitado.')
        adjuntar_pdf = data.get('adjuntar_pdf', True)
        adjuntar_excel = data.get('adjuntar_excel', True)
        
        email = EmailMessage(
            subject=subject,
            body=body,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[destinatario],
        )
        
        # 1. Adjuntar PDF si es solicitado
        if adjuntar_pdf and pdf_template:
            html_string = render_to_string(pdf_template, pdf_context, request=request)
            pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()
            email.attach(pdf_filename, pdf_file, 'application/pdf')
            
        # 2. Adjuntar Excel si es solicitado y existe función generadora
        if adjuntar_excel and write_excel_data_fn:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Reporte'
            # Escribir los datos usando la función de callback
            write_excel_data_fn(ws)
            # Aplicar formato de tabla nativa y estilos
            _estilo_excel_sheet(ws, excel_headers)
            
            # Guardar en buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            email.attach(excel_filename, buffer.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            
        email.send()
        return JsonResponse({'status': 'success', 'message': f'Reporte enviado a {destinatario}.'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'error': f'Error al enviar correo: {str(e)}'}, status=500)


# ══════════════════════════════════════════
#  TIPOS DE SUSCRIPCIÓN
# ══════════════════════════════════════════

MONEDAS = ['USD', 'EUR', 'GBP', 'AUD', 'BRL', 'CAD', 'CHF',
           'CNY', 'INR', 'JPY', 'KRW', 'MXN', 'RUB', 'ZAR']


@login_required
def plan_list(request):
    query = request.GET.get('q', '')
    filtro = {}
    if query:
        filtro = {'$or': [
            {'nombrePlan': {'$regex': query, '$options': 'i'}},
            {'moneda':     {'$regex': query, '$options': 'i'}},
        ]}
    planes = [MongoDoc(d) for d in tipo_suscripciones_col.find(filtro).sort('tipo_id', 1)]
    import sys
    if planes:
        print(f"DEBUG VIEW: plan type is {type(planes[0])}", file=sys.stderr)
        print(f"DEBUG VIEW: plan._data is {planes[0]._data}", file=sys.stderr)
        print(f"DEBUG VIEW: plan.tipo_id evaluates to {getattr(planes[0], 'tipo_id', 'NOT FOUND')}", file=sys.stderr)
    return render(request, 'Suscripciones/plan/plan_list.html', {
        'planes': planes,
        'query': query,
    })


@login_required
def plan_add(request):
    if request.method == 'POST':
        try:
            nombre  = request.POST.get('nombreplan')
            precio  = request.POST.get('precio')
            moneda  = request.POST.get('moneda')
            duracion = request.POST.get('duracion')

            if not all([nombre, precio, moneda, duracion]):
                messages.error(request, 'Todos los campos son obligatorios.')
                return render(request, 'Suscripciones/plan/plan_form.html',
                              {'action': 'Nuevo', 'monedas': MONEDAS})

            tipo_suscripciones_col.insert_one({
                'tipo_id':    _next_id(tipo_suscripciones_col, 'tipo_id'),
                'nombrePlan': nombre,
                'precio':     float(precio),
                'moneda':     moneda,
                'duracion':   int(duracion),
            })
            messages.success(request, f"Plan '{nombre}' creado correctamente.")
            return redirect('plan_list')
        except Exception as e:
            messages.error(request, f'Error al crear el plan: {e}')
    return render(request, 'Suscripciones/plan/plan_form.html',
                  {'action': 'Nuevo', 'monedas': MONEDAS})


@login_required
def plan_edit(request, pk):
    doc = _get_or_404(tipo_suscripciones_col, {'tipo_id': pk}, request, 'plan_list')
    if not doc:
        return redirect('plan_list')

    if request.method == 'POST':
        try:
            new_name = request.POST.get('nombreplan')
            tipo_suscripciones_col.update_one({'tipo_id': pk}, {'$set': {
                'nombrePlan': new_name,
                'precio':     float(request.POST.get('precio')),
                'moneda':     request.POST.get('moneda'),
                'duracion':   int(request.POST.get('duracion')),
            }})
            messages.success(request, f"Plan '{new_name}' actualizado.")
            return redirect('plan_list')
        except Exception as e:
            messages.error(request, f'Error al actualizar: {e}')
            doc = tipo_suscripciones_col.find_one({'tipo_id': pk})

    return render(request, 'Suscripciones/plan/plan_form.html', {
        'action': 'Editar',
        'plan':    MongoDoc(doc),
        'monedas': MONEDAS,
    })


@login_required
def plan_delete(request, pk):
    doc = _get_or_404(tipo_suscripciones_col, {'tipo_id': pk}, request, 'plan_list')
    if not doc:
        return redirect('plan_list')

    if request.method == 'POST':
        nombre = doc.get('nombrePlan')
        tipo_suscripciones_col.delete_one({'tipo_id': pk})
        messages.success(request, f"Plan '{nombre}' eliminado.")
        return redirect('plan_list')

    return render(request, 'Suscripciones/confirm_delete.html', {
        'objeto':     doc.get('nombrePlan'),
        'tipo':       'plan de suscripción',
        'cancel_url': 'plan_list',
    })


# ══════════════════════════════════════════
#  PROMOCIONES
# ══════════════════════════════════════════

@login_required
def promocion_list(request):
    query = request.GET.get('q', '')
    filtro = {}
    if query:
        filtro = {'$or': [
            {'descripcion':                {'$regex': query, '$options': 'i'}},
            {'tipoSuscripcion.nombrePlan': {'$regex': query, '$options': 'i'}},
        ]}
    docs = list(promociones_col.find(filtro).sort('promo_id', 1))
    for d in docs:
        prepare_doc(d)
    promociones = [MongoDoc(d) for d in docs]
    return render(request, 'Suscripciones/promocion/promocion_list.html', {
        'promociones': promociones,
        'query': query,
    })


@login_required
def promocion_add(request):
    planes = _planes_dropdown()
    if request.method == 'POST':
        try:
            descripcion    = request.POST.get('descripcion')
            porcentajedesc = request.POST.get('porcentajedesc')
            fechainicio     = request.POST.get('fechainicio')
            fechaexpira     = request.POST.get('fechaexpira') or None
            estadoactivo    = request.POST.get('estadoactivo') == 'on'
            idtipo          = request.POST.get('tiposuscripcion')

            if not all([descripcion, porcentajedesc, fechainicio, idtipo]):
                messages.error(request, 'Faltan campos obligatorios.')
                return render(request, 'Suscripciones/promocion/promocion_form.html',
                              {'action': 'Nueva', 'planes': planes})

            # Buscar el plan para embeber
            plan_doc = tipo_suscripciones_col.find_one({'tipo_id': int(idtipo)})
            tipo_embed = {
                '_id':        int(idtipo),
                'nombrePlan': plan_doc['nombrePlan'] if plan_doc else '',
            }

            fecha_inicio_str = f"{fechainicio}T00:00:00" if fechainicio and 'T' not in fechainicio else fechainicio
            fecha_expira_str = (f"{fechaexpira}T00:00:00" if fechaexpira and 'T' not in fechaexpira else fechaexpira) if fechaexpira else None

            promociones_col.insert_one({
                'promo_id':         _next_id(promociones_col, 'promo_id'),
                'descripcion':      descripcion,
                'porcentajeDesc':   int(porcentajedesc),
                'fechaInicio':      fecha_inicio_str,
                'fechaExpira':      fecha_expira_str,
                'estadoActivo':     estadoactivo,
                'tipoSuscripcion':  tipo_embed,
            })
            messages.success(request, f"Promoción '{descripcion}' creada.")
            return redirect('promocion_list')
        except Exception as e:
            messages.error(request, f'Error: {e}')
    return render(request, 'Suscripciones/promocion/promocion_form.html',
                  {'action': 'Nueva', 'planes': planes})


@login_required
def promocion_edit(request, pk):
    doc = _get_or_404(promociones_col, {'promo_id': pk}, request, 'promocion_list')
    if not doc:
        return redirect('promocion_list')

    planes = _planes_dropdown()

    if request.method == 'POST':
        try:
            desc     = request.POST.get('descripcion')
            idtipo   = request.POST.get('tiposuscripcion')
            fechainicio = request.POST.get('fechainicio')
            fechaexpira = request.POST.get('fechaexpira') or None

            plan_doc = tipo_suscripciones_col.find_one({'tipo_id': int(idtipo)})
            tipo_embed = {
                '_id':        int(idtipo),
                'nombrePlan': plan_doc['nombrePlan'] if plan_doc else '',
            }

            fecha_inicio_str = f"{fechainicio}T00:00:00" if fechainicio and 'T' not in fechainicio else fechainicio
            fecha_expira_str = (f"{fechaexpira}T00:00:00" if fechaexpira and 'T' not in fechaexpira else fechaexpira) if fechaexpira else None

            promociones_col.update_one({'promo_id': pk}, {'$set': {
                'descripcion':     desc,
                'porcentajeDesc':  int(request.POST.get('porcentajedesc')),
                'fechaInicio':     fecha_inicio_str,
                'fechaExpira':     fecha_expira_str,
                'estadoActivo':    request.POST.get('estadoactivo') == 'on',
                'tipoSuscripcion': tipo_embed,
            }})
            messages.success(request, f"Promoción '{desc}' actualizada.")
            return redirect('promocion_list')
        except Exception as e:
            messages.error(request, f'Error: {e}')
            doc = promociones_col.find_one({'promo_id': pk})

    prepare_doc(doc)
    # Añadir campo auxiliar para el select del formulario
    doc['tiposuscripcion_id'] = doc.get('tipoSuscripcion', {}).get('_id')
    return render(request, 'Suscripciones/promocion/promocion_form.html', {
        'action':    'Editar',
        'promocion': MongoDoc(doc),
        'planes':    planes,
    })


@login_required
def promocion_delete(request, pk):
    doc = _get_or_404(promociones_col, {'promo_id': pk}, request, 'promocion_list')
    if not doc:
        return redirect('promocion_list')

    if request.method == 'POST':
        desc = doc.get('descripcion')
        promociones_col.delete_one({'promo_id': pk})
        messages.success(request, f"Promoción '{desc}' eliminada.")
        return redirect('promocion_list')

    return render(request, 'Suscripciones/confirm_delete.html', {
        'objeto':     doc.get('descripcion'),
        'tipo':       'promoción',
        'cancel_url': 'promocion_list',
    })


# ══════════════════════════════════════════
#  SUSCRIPCIONES
# ══════════════════════════════════════════

ESTADOS_SUSCRIPCION = ['Activa', 'Cancelada', 'Expirada']


@login_required
def suscripcion_list(request):
    query = request.GET.get('q', '')
    filtro = {}
    if query:
        filtro = {'$or': [
            {'estado':                       {'$regex': query, '$options': 'i'}},
            {'tipoSuscripcion.nombrePlan':   {'$regex': query, '$options': 'i'}},
        ]}

    docs = list(suscripciones_col.find(filtro))
    # Lookup nombres de usuario
    user_ids = {d.get('idUsuario') for d in docs if d.get('idUsuario')}
    user_names = _lookup_usuarios(user_ids)

    for d in docs:
        prepare_doc(d)
        uid = d.get('idUsuario')
        d['usuario_nombre'] = user_names.get(uid, '—')

    suscripciones = [MongoDoc(d) for d in docs]
    return render(request, 'Suscripciones/suscripcion/suscripcion_list.html', {
        'suscripciones': suscripciones,
        'query': query,
    })


@login_required
def suscripcion_add(request):
    planes      = _planes_dropdown()
    promociones = _promociones_dropdown(solo_activas=True)
    usuarios    = _usuarios_dropdown()

    if request.method == 'POST':
        try:
            fechainicio = request.POST.get('fechainicio')
            fechafin    = request.POST.get('fechafin')
            estado      = request.POST.get('estado')
            idusuario   = request.POST.get('usuario') or None
            idtipo      = request.POST.get('tiposuscripcion')
            idpromo     = request.POST.get('promocion') or None

            if not all([fechainicio, fechafin, estado, idtipo]):
                messages.error(request, 'Faltan campos obligatorios.')
                return render(request, 'Suscripciones/suscripcion/suscripcion_form.html', {
                    'action': 'Nueva', 'planes': planes,
                    'promociones': promociones, 'usuarios': usuarios,
                    'estados': ESTADOS_SUSCRIPCION,
                })

            # Embeber tipoSuscripcion
            plan_doc = tipo_suscripciones_col.find_one({'tipo_id': int(idtipo)})
            tipo_embed = {
                '_id':        int(idtipo),
                'nombrePlan': plan_doc['nombrePlan'] if plan_doc else '',
                'precio':     plan_doc['precio'] if plan_doc else 0,
            }

            # Embeber promocion
            promo_embed = {'_id': None, 'descripcion': None, 'porcentajeDesc': None}
            if idpromo:
                promo_doc = promociones_col.find_one({'promo_id': int(idpromo)})
                if promo_doc:
                    promo_embed = {
                        '_id':             int(idpromo),
                        'descripcion':     promo_doc.get('descripcion'),
                        'porcentajeDesc':  promo_doc.get('porcentajeDesc'),
                    }

            fi = f"{fechainicio}T00:00:00" if fechainicio and 'T' not in fechainicio else fechainicio
            ff = f"{fechafin}T00:00:00" if fechafin and 'T' not in fechafin else fechafin

            suscripciones_col.insert_one({
                'fechaInicio':      fi,
                'fechaFin':         ff,
                'estado':           estado,
                'idUsuario':        int(idusuario) if idusuario else None,
                'tipoSuscripcion':  tipo_embed,
                'promocion':        promo_embed,
            })
            messages.success(request, 'Suscripción registrada correctamente.')
            return redirect('suscripcion_list')
        except Exception as e:
            messages.error(request, f'Error: {e}')

    return render(request, 'Suscripciones/suscripcion/suscripcion_form.html', {
        'action':      'Nueva',
        'planes':      planes,
        'promociones': promociones,
        'usuarios':    usuarios,
        'estados':     ESTADOS_SUSCRIPCION,
    })


@login_required
def suscripcion_edit(request, pk):
    doc = _get_or_404(suscripciones_col, {'_id': ObjectId(pk)}, request, 'suscripcion_list')
    if not doc:
        return redirect('suscripcion_list')

    planes      = _planes_dropdown()
    promociones = _promociones_dropdown(solo_activas=True)
    usuarios    = _usuarios_dropdown()

    if request.method == 'POST':
        try:
            fechainicio = request.POST.get('fechainicio')
            fechafin    = request.POST.get('fechafin')
            estado      = request.POST.get('estado')
            idusuario   = request.POST.get('usuario') or None
            idtipo      = request.POST.get('tiposuscripcion')
            idpromo     = request.POST.get('promocion') or None

            plan_doc = tipo_suscripciones_col.find_one({'tipo_id': int(idtipo)})
            tipo_embed = {
                '_id':        int(idtipo),
                'nombrePlan': plan_doc['nombrePlan'] if plan_doc else '',
                'precio':     plan_doc['precio'] if plan_doc else 0,
            }

            promo_embed = {'_id': None, 'descripcion': None, 'porcentajeDesc': None}
            if idpromo:
                promo_doc = promociones_col.find_one({'promo_id': int(idpromo)})
                if promo_doc:
                    promo_embed = {
                        '_id':             int(idpromo),
                        'descripcion':     promo_doc.get('descripcion'),
                        'porcentajeDesc':  promo_doc.get('porcentajeDesc'),
                    }

            fi = f"{fechainicio}T00:00:00" if fechainicio and 'T' not in fechainicio else fechainicio
            ff = f"{fechafin}T00:00:00" if fechafin and 'T' not in fechafin else fechafin

            suscripciones_col.update_one({'_id': ObjectId(pk)}, {'$set': {
                'fechaInicio':      fi,
                'fechaFin':         ff,
                'estado':           estado,
                'idUsuario':        int(idusuario) if idusuario else None,
                'tipoSuscripcion':  tipo_embed,
                'promocion':        promo_embed,
            }})
            messages.success(request, 'Suscripción actualizada.')
            return redirect('suscripcion_list')
        except Exception as e:
            messages.error(request, f'Error: {e}')
            doc = suscripciones_col.find_one({'_id': ObjectId(pk)})

    prepare_doc(doc)
    # Campos auxiliares para los selects del formulario
    doc['tiposuscripcion_id'] = doc.get('tipoSuscripcion', {}).get('_id')
    promo = doc.get('promocion', {})
    doc['promocion_id'] = promo.get('_id') if promo else None
    doc['usuario_id'] = doc.get('idUsuario')

    return render(request, 'Suscripciones/suscripcion/suscripcion_form.html', {
        'action':       'Editar',
        'suscripcion':  MongoDoc(doc),
        'planes':       planes,
        'promociones':  promociones,
        'usuarios':     usuarios,
        'estados':      ESTADOS_SUSCRIPCION,
    })


@login_required
def suscripcion_delete(request, pk):
    doc = _get_or_404(suscripciones_col, {'_id': ObjectId(pk)}, request, 'suscripcion_list')
    if not doc:
        return redirect('suscripcion_list')

    if request.method == 'POST':
        suscripciones_col.delete_one({'_id': ObjectId(pk)})
        messages.success(request, 'Suscripción eliminada.')
        return redirect('suscripcion_list')

    plan_name = doc.get('tipoSuscripcion', {}).get('nombrePlan', '')
    return render(request, 'Suscripciones/confirm_delete.html', {
        'objeto':     f"Suscripción {plan_name} — {doc.get('estado', '')}",
        'tipo':       'suscripción',
        'cancel_url': 'suscripcion_list',
    })


# ══════════════════════════════════════════
#  NOTIFICACIONES
# ══════════════════════════════════════════

TIPOS_NOTIFICACION = ['Aviso de Seguridad', 'Nuevo Lanzamiento', 'Pago Exitoso']


@login_required
def notificacion_list(request):
    query = request.GET.get('q', '')
    filtro = {}
    if query:
        filtro = {'$or': [
            {'tipoNotif': {'$regex': query, '$options': 'i'}},
            {'mensaje':   {'$regex': query, '$options': 'i'}},
        ]}

    docs = list(notificaciones_col.find(filtro))

    # Lookup usuarios y promociones
    user_ids  = {d.get('idUsuario') for d in docs if d.get('idUsuario')}
    promo_ids = {d.get('idPromocion') for d in docs if d.get('idPromocion')}
    user_names = _lookup_usuarios(user_ids)
    promo_descs = {}
    if promo_ids:
        for p in promociones_col.find({'promo_id': {'$in': list(promo_ids)}}):
            promo_descs[p['promo_id']] = p.get('descripcion', '')

    for d in docs:
        prepare_doc(d)
        d['usuario_nombre'] = user_names.get(d.get('idUsuario'), '—')
        d['promocion_desc'] = promo_descs.get(d.get('idPromocion'), '—')

    notificaciones = [MongoDoc(d) for d in docs]
    return render(request, 'Suscripciones/notificacion/notificacion_list.html', {
        'notificaciones': notificaciones,
        'query': query,
    })


@login_required
def notificacion_add(request):
    usuarios    = _usuarios_dropdown()
    promociones = _promociones_dropdown()

    if request.method == 'POST':
        try:
            tiponotif = request.POST.get('tiponotif')
            mensaje   = request.POST.get('mensaje')
            idusuario = request.POST.get('usuario')
            idpromo   = request.POST.get('promocion')

            if not all([tiponotif, mensaje, idusuario, idpromo]):
                messages.error(request, 'Todos los campos son obligatorios.')
                return render(request, 'Suscripciones/notificacion/notificacion_form.html', {
                    'action': 'Nueva', 'usuarios': usuarios,
                    'promociones': promociones, 'tipos': TIPOS_NOTIFICACION,
                })

            notificaciones_col.insert_one({
                'tipoNotif':   tiponotif,
                'mensaje':     mensaje,
                'fechaEnvio':  datetime.now().isoformat(),
                'idUsuario':   int(idusuario),
                'idPromocion': int(idpromo),
            })
            messages.success(request, 'Notificación creada correctamente.')
            return redirect('notificacion_list')
        except Exception as e:
            messages.error(request, f'Error: {e}')

    return render(request, 'Suscripciones/notificacion/notificacion_form.html', {
        'action':      'Nueva',
        'usuarios':    usuarios,
        'promociones': promociones,
        'tipos':       TIPOS_NOTIFICACION,
    })


@login_required
def notificacion_delete(request, pk):
    doc = _get_or_404(notificaciones_col, {'_id': ObjectId(pk)}, request, 'notificacion_list')
    if not doc:
        return redirect('notificacion_list')

    if request.method == 'POST':
        notificaciones_col.delete_one({'_id': ObjectId(pk)})
        messages.success(request, 'Notificación eliminada.')
        return redirect('notificacion_list')

    return render(request, 'Suscripciones/confirm_delete.html', {
        'objeto':     f"{doc.get('tipoNotif', '')} — {doc.get('mensaje', '')[:50]}",
        'tipo':       'notificación',
        'cancel_url': 'notificacion_list',
    })


# ══════════════════════════════════════════
#  PLAYLISTS
# ══════════════════════════════════════════

@login_required
def playlist_list(request):
    query = request.GET.get('q', '')
    filtro = {}
    if query:
        filtro = {'$or': [
            {'nombre': {'$regex': query, '$options': 'i'}},
        ]}

    docs = list(playlists_col.find(filtro))
    user_ids = {d.get('idUsuario') for d in docs if d.get('idUsuario')}
    user_names = _lookup_usuarios(user_ids)

    for d in docs:
        prepare_doc(d)
        d['usuario_nombre'] = user_names.get(d.get('idUsuario'), '—')

    playlists = [MongoDoc(d) for d in docs]
    return render(request, 'Suscripciones/playlist/playlist_list.html', {
        'playlists': playlists,
        'query': query,
    })


@login_required
def playlist_add(request):
    usuarios = _usuarios_dropdown()

    if request.method == 'POST':
        try:
            nombre = request.POST.get('nombre')
            if not nombre:
                messages.error(request, 'El nombre es obligatorio.')
                return render(request, 'Suscripciones/playlist/playlist_form.html',
                              {'action': 'Nueva', 'usuarios': usuarios})

            idusuario = request.POST.get('usuario') or None
            playlists_col.insert_one({
                'nombre':        nombre,
                'descripcion':   request.POST.get('descripcion') or '',
                'esPrivada':     request.POST.get('esprivada') == 'on',
                'esPublicada':   request.POST.get('espublicada') == 'on',
                'imagenPortada': request.POST.get('imagenportada') or '',
                'fechaCreacion': datetime.now().isoformat(),
                'idUsuario':     int(idusuario) if idusuario else None,
                'canciones':     [],
            })
            messages.success(request, f"Playlist '{nombre}' creada.")
            return redirect('playlist_list')
        except Exception as e:
            messages.error(request, f'Error: {e}')

    return render(request, 'Suscripciones/playlist/playlist_form.html', {
        'action':   'Nueva',
        'usuarios': usuarios,
    })


@login_required
def playlist_edit(request, pk):
    doc = _get_or_404(playlists_col, {'_id': ObjectId(pk)}, request, 'playlist_list')
    if not doc:
        return redirect('playlist_list')

    usuarios = _usuarios_dropdown()

    if request.method == 'POST':
        try:
            idusuario = request.POST.get('usuario') or None
            playlists_col.update_one({'_id': ObjectId(pk)}, {'$set': {
                'nombre':        request.POST.get('nombre'),
                'descripcion':   request.POST.get('descripcion') or '',
                'esPrivada':     request.POST.get('esprivada') == 'on',
                'esPublicada':   request.POST.get('espublicada') == 'on',
                'imagenPortada': request.POST.get('imagenportada') or '',
                'idUsuario':     int(idusuario) if idusuario else None,
            }})
            messages.success(request, f"Playlist actualizada.")
            return redirect('playlist_list')
        except Exception as e:
            messages.error(request, f'Error: {e}')
            doc = playlists_col.find_one({'_id': ObjectId(pk)})

    prepare_doc(doc)
    doc['usuario_id'] = doc.get('idUsuario')
    return render(request, 'Suscripciones/playlist/playlist_form.html', {
        'action':   'Editar',
        'playlist': MongoDoc(doc),
        'usuarios': usuarios,
    })


@login_required
def playlist_delete(request, pk):
    doc = _get_or_404(playlists_col, {'_id': ObjectId(pk)}, request, 'playlist_list')
    if not doc:
        return redirect('playlist_list')

    if request.method == 'POST':
        nombre = doc.get('nombre')
        playlists_col.delete_one({'_id': ObjectId(pk)})
        messages.success(request, f"Playlist '{nombre}' eliminada.")
        return redirect('playlist_list')

    return render(request, 'Suscripciones/confirm_delete.html', {
        'objeto':     doc.get('nombre'),
        'tipo':       'playlist',
        'cancel_url': 'playlist_list',
    })


# ══════════════════════════════════════════
#  PLAYLIST — CANCIONES (embebidas)
# ══════════════════════════════════════════

@login_required
def playlist_canciones(request, pk):
    doc = _get_or_404(playlists_col, {'_id': ObjectId(pk)}, request, 'playlist_list')
    if not doc:
        return redirect('playlist_list')

    prepare_doc(doc)
    canciones_data = doc.get('canciones') or []

    # Lookup nombres de canciones
    cancion_ids = [c.get('idCancion') for c in canciones_data if c.get('idCancion')]
    canciones_lookup = {}
    if cancion_ids:
        for c in canciones_col.find({'cancion_id': {'$in': cancion_ids}}):
            canciones_lookup[c['cancion_id']] = c.get('tituloCancion', f"Canción #{c['cancion_id']}")

    entradas = []
    for c in canciones_data:
        c['cancion_nombre'] = canciones_lookup.get(c.get('idCancion'), f"Canción #{c.get('idCancion', '?')}")
        if isinstance(c.get('fechaAdicion'), str):
            c['fechaAdicion'] = parse_date(c['fechaAdicion'])
        entradas.append(MongoDoc(c))

    # Ordenar por el campo 'orden'
    entradas.sort(key=lambda e: e.orden or 0)

    playlist = MongoDoc(doc)
    return render(request, 'Suscripciones/playlist/playlist_canciones.html', {
        'playlist': playlist,
        'entradas': entradas,
    })


@login_required
def playlist_cancion_agregar(request, pk):
    doc = _get_or_404(playlists_col, {'_id': ObjectId(pk)}, request, 'playlist_list')
    if not doc:
        return redirect('playlist_list')

    canciones = _canciones_dropdown()

    if request.method == 'POST':
        try:
            cancion_id = int(request.POST.get('cancion'))
            orden      = int(request.POST.get('orden') or 0)

            # Verificar duplicados
            existing = doc.get('canciones') or []
            if any(c.get('idCancion') == cancion_id for c in existing):
                messages.error(request, 'Esa canción ya está en la playlist.')
                prepare_doc(doc)
                return render(request, 'Suscripciones/playlist/playlist_cancion_form.html', {
                    'playlist': MongoDoc(doc), 'canciones': canciones,
                })

            playlists_col.update_one({'_id': ObjectId(pk)}, {'$push': {
                'canciones': {
                    'idCancion':    cancion_id,
                    'fechaAdicion': datetime.now().isoformat(),
                    'orden':        orden,
                }
            }})
            messages.success(request, 'Canción agregada a la playlist.')
            return redirect('playlist_canciones', pk=pk)
        except Exception as e:
            messages.error(request, f'Error: {e}')

    prepare_doc(doc)
    return render(request, 'Suscripciones/playlist/playlist_cancion_form.html', {
        'playlist':  MongoDoc(doc),
        'canciones': canciones,
    })


@login_required
def playlist_cancion_quitar(request, pk):
    if request.method == 'POST':
        cancion_id = int(request.POST.get('cancion_id', 0))
        result = playlists_col.update_one(
            {'_id': ObjectId(pk)},
            {'$pull': {'canciones': {'idCancion': cancion_id}}}
        )
        if result.modified_count:
            messages.success(request, 'Canción quitada de la playlist.')
        else:
            messages.error(request, 'No se encontró esa canción en la playlist.')

    return redirect('playlist_canciones', pk=pk)


# ══════════════════════════════════════════
#  ESTADÍSTICAS DIARIAS
# ══════════════════════════════════════════

@login_required
def estadistica_list(request):
    query = request.GET.get('q', '')
    filtro = {}
    # La búsqueda por nombre de canción requiere lookup
    docs = list(estadisticas_col.find().sort('fechaReporte', -1))

    # Lookup canciones
    cancion_ids = list({d.get('idCancion') for d in docs if d.get('idCancion')})
    canciones_lookup = {}
    if cancion_ids:
        for c in canciones_col.find({'cancion_id': {'$in': cancion_ids}}):
            canciones_lookup[c['cancion_id']] = c

    filtered = []
    for d in docs:
        cid = d.get('idCancion')
        cancion = canciones_lookup.get(cid, {})
        d['cancion_nombre']  = cancion.get('tituloCancion', f'Canción #{cid}')
        d['cancion_portada'] = cancion.get('urlPortada', '')
        d['cancion_album']   = cancion.get('album_id', '')
        prepare_doc(d)

        # Filtro de búsqueda (se hace en Python porque depende del lookup)
        if query:
            search = query.lower()
            if (search not in d['cancion_nombre'].lower()
                    and search not in str(d.get('fechaReporte', ''))):
                continue
        filtered.append(d)

    estadisticas = [MongoDoc(d) for d in filtered]
    return render(request, 'Suscripciones/estadistica/estadistica_list.html', {
        'estadisticas': estadisticas,
        'query': query,
    })


@login_required
def estadistica_add(request):
    canciones = _canciones_dropdown()

    if request.method == 'POST':
        try:
            totalrepros  = request.POST.get('totalrepros')
            fechareporte = request.POST.get('fechareporte')
            cancion_id   = request.POST.get('cancion')

            if not all([totalrepros, fechareporte, cancion_id]):
                messages.error(request, 'Todos los campos son obligatorios.')
                return render(request, 'Suscripciones/estadistica/estadistica_form.html',
                              {'action': 'Nueva', 'canciones': canciones})

            estadisticas_col.insert_one({
                'totalRepros':  int(totalrepros),
                'fechaReporte': fechareporte,
                'idCancion':    int(cancion_id),
            })
            messages.success(request, 'Estadística registrada.')
            return redirect('estadistica_list')
        except Exception as e:
            messages.error(request, f'Error: {e}')

    return render(request, 'Suscripciones/estadistica/estadistica_form.html', {
        'action':    'Nueva',
        'canciones': canciones,
    })


@login_required
def estadistica_delete(request, pk):
    doc = _get_or_404(estadisticas_col, {'_id': ObjectId(pk)}, request, 'estadistica_list')
    if not doc:
        return redirect('estadistica_list')

    if request.method == 'POST':
        estadisticas_col.delete_one({'_id': ObjectId(pk)})
        messages.success(request, 'Estadística eliminada.')
        return redirect('estadistica_list')

    return render(request, 'Suscripciones/confirm_delete.html', {
        'objeto':     f"Estadística del {doc.get('fechaReporte', '')}",
        'tipo':       'estadística',
        'cancel_url': 'estadistica_list',
    })


# ══════════════════════════════════════════
#  REPORTES
# ══════════════════════════════════════════

@login_required
def reporte_vencimientos(request):
    """Informe B: Suscripciones que vencen en los próximos 5 días."""
    desde, hasta, desde_str, hasta_str = _parse_date_params(request)
    docs = _data_vencimientos(desde, hasta)
    docs_paginated, total, page, total_pages, per_page = _paginate(request, docs)
    return render(request, 'Suscripciones/reportes/reporte_vencimientos.html', {
        'proximas': docs_paginated,
        'hoy':      datetime.now(),
        'limite':   datetime.now() + timedelta(days=5),
        'desde': desde_str,
        'hasta': hasta_str,
        'page': page,
        'total_pages': total_pages,
        'total': total,
        'per_page': per_page,
    })


@login_required
def reporte_promociones_vencidas(request):
    """Informe D: Promociones expiradas que aún figuran como activas."""
    desde, hasta, desde_str, hasta_str = _parse_date_params(request)
    docs = _data_promociones_vencidas(desde, hasta)
    docs_paginated, total, page, total_pages, per_page = _paginate(request, docs)
    return render(request, 'Suscripciones/reportes/reporte_promociones_vencidas.html', {
        'vencidas': docs_paginated,
        'ahora':    datetime.now(),
        'desde': desde_str,
        'hasta': hasta_str,
        'page': page,
        'total_pages': total_pages,
        'total': total,
        'per_page': per_page,
    })


# Helper to get base context for reports
def _contexto_base(request):
    return {
        'fecha_generacion': datetime.now(),
        'usuario_generador': (f"{request.user.get_full_name() or request.user.username} · {request.user.email}").strip(' ·'),
        'version': 'v2.1.0',
    }

# ── EXPORTACIÓN SUSCRIPCIONES EXPIRADAS (VENCIMIENTOS) ──
@login_required
def exportar_vencimientos_pdf(request):
    try:
        docs = _data_vencimientos()
        context = {**_contexto_base(request), 'proximas': docs, 'hoy': datetime.now(), 'limite': datetime.now() + timedelta(days=5)}
        return _render_pdf(request, 'Suscripciones/reportes/pdf_vencimientos.html', context, 'Suscripciones_Expiradas.pdf')
    except Exception as e:
        return HttpResponse(f"Error: {str(e)}", status=500)

@login_required
def enviar_vencimientos_correo(request):
    try:
        docs = _data_vencimientos()
        pdf_context = {**_contexto_base(request), 'proximas': docs, 'hoy': datetime.now(), 'limite': datetime.now() + timedelta(days=5)}
        headers = ['Usuario', 'Plan', 'Fecha Inicio', 'Fecha Fin', 'Estado']
        
        def write_excel_data(ws):
            for i, d in enumerate(docs, 2):
                ws.cell(row=i, column=1, value=d.usuario_nombre or '—')
                nombre_plan = d.tipoSuscripcion.nombrePlan if d.tipoSuscripcion else '—'
                ws.cell(row=i, column=2, value=nombre_plan)
                ws.cell(row=i, column=3, value=str(d.fechaInicio or '—'))
                ws.cell(row=i, column=4, value=str(d.fechaFin or '—'))
                ws.cell(row=i, column=5, value=d.estado or '—')
                
        return _dispatch_report_email(
            request=request,
            pdf_template='Suscripciones/reportes/pdf_vencimientos.html',
            pdf_context=pdf_context,
            pdf_filename='Suscripciones_Expiradas.pdf',
            excel_headers=headers,
            write_excel_data_fn=write_excel_data,
            excel_filename='Suscripciones_Expiradas.xlsx'
        )
    except Exception as e:
        return JsonResponse({'status': 'error', 'error': str(e)}, status=500)

@login_required
def exportar_vencimientos_excel(request):
    try:
        import openpyxl
        docs = _data_vencimientos()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Vencimientos'
        headers = ['Usuario', 'Plan', 'Fecha Inicio', 'Fecha Fin', 'Estado']
        for i, d in enumerate(docs, 2):
            ws.cell(row=i, column=1, value=d.usuario_nombre or '—')
            nombre_plan = d.tipoSuscripcion.nombrePlan if d.tipoSuscripcion else '—'
            ws.cell(row=i, column=2, value=nombre_plan)
            ws.cell(row=i, column=3, value=str(d.fechaInicio or '—'))
            ws.cell(row=i, column=4, value=str(d.fechaFin or '—'))
            ws.cell(row=i, column=5, value=d.estado or '—')
        _estilo_excel_sheet(ws, headers)
        return _build_excel_response(wb, 'Suscripciones_Expiradas.xlsx')
    except Exception as e:
        return HttpResponse(f"Error: {str(e)}", status=500)


# ── EXPORTACIÓN PROMOCIONES ALTO DESCUENTO ──
@login_required
def exportar_promociones_vencidas_pdf(request):
    try:
        docs = _data_promociones_vencidas()
        context = {**_contexto_base(request), 'vencidas': docs, 'ahora': datetime.now()}
        return _render_pdf(request, 'Suscripciones/reportes/pdf_promociones_vencidas.html', context, 'Promociones_Alto_Descuento.pdf')
    except Exception as e:
        return HttpResponse(f"Error: {str(e)}", status=500)

@login_required
def enviar_promociones_vencidas_correo(request):
    try:
        docs = _data_promociones_vencidas()
        pdf_context = {**_contexto_base(request), 'vencidas': docs, 'ahora': datetime.now()}
        headers = ['Promoción ID', 'Nombre', 'Descuento %', 'Fecha Expira', 'Estado']
        
        def write_excel_data(ws):
            for i, d in enumerate(docs, 2):
                ws.cell(row=i, column=1, value=d.promo_id or '—')
                ws.cell(row=i, column=2, value=d.nombre or '—')
                ws.cell(row=i, column=3, value=d.porcentajeDesc or 0)
                ws.cell(row=i, column=4, value=str(d.fechaExpira or '—'))
                ws.cell(row=i, column=5, value='Activa' if d.estadoActivo else 'Inactiva')
                
        return _dispatch_report_email(
            request=request,
            pdf_template='Suscripciones/reportes/pdf_promociones_vencidas.html',
            pdf_context=pdf_context,
            pdf_filename='Promociones_Alto_Descuento.pdf',
            excel_headers=headers,
            write_excel_data_fn=write_excel_data,
            excel_filename='Promociones_Alto_Descuento.xlsx'
        )
    except Exception as e:
        return JsonResponse({'status': 'error', 'error': str(e)}, status=500)

@login_required
def exportar_promociones_vencidas_excel(request):
    try:
        import openpyxl
        docs = _data_promociones_vencidas()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Promociones'
        headers = ['Promoción ID', 'Nombre', 'Descuento %', 'Fecha Expira', 'Estado']
        for i, d in enumerate(docs, 2):
            ws.cell(row=i, column=1, value=d.promo_id or '—')
            ws.cell(row=i, column=2, value=d.nombre or '—')
            ws.cell(row=i, column=3, value=d.porcentajeDesc or 0)
            ws.cell(row=i, column=4, value=str(d.fechaExpira or '—'))
            ws.cell(row=i, column=5, value='Activa' if d.estadoActivo else 'Inactiva')
        _estilo_excel_sheet(ws, headers)
        return _build_excel_response(wb, 'Promociones_Alto_Descuento.xlsx')
    except Exception as e:
        return HttpResponse(f"Error: {str(e)}", status=500)


# ── REPORTE: USUARIOS PREMIUM ACTIVOS ──
def obtener_usuarios_premium_activos():
    users = list(db["usuarios"].find({
        "estadoPlan": "Premium",
        "estadoActivo": "Activo"
    }))
    for u in users:
        u["id"] = str(u.get("_id"))
        u["nombre_completo"] = f"{u.get('nombre', '')} {u.get('apellido', '')}"
        if isinstance(u.get("fechaRegistro"), str):
            u["fechaRegistro"] = parse_date(u["fechaRegistro"])
    return users

@login_required
def reporte_usuarios_premium(request):
    try:
        usuarios = obtener_usuarios_premium_activos()
        usuarios, total, page, total_pages, per_page = _paginate(request, usuarios)
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")
        usuarios = []
        total = page = total_pages = per_page = 0
    return render(request, 'Suscripciones/reportes/reporte_usuarios_premium.html', {
        'usuarios': usuarios, 'page': page, 'total_pages': total_pages, 'total': total, 'per_page': per_page,
    })

@login_required
def exportar_usuarios_premium_pdf(request):
    try:
        usuarios = obtener_usuarios_premium_activos()
        context = {**_contexto_base(request), 'usuarios': usuarios}
        return _render_pdf(request, 'Suscripciones/reportes/pdf_usuarios_premium.html', context, 'Usuarios_Premium_Activos.pdf')
    except Exception as e:
        return HttpResponse(f"Error: {str(e)}", status=500)

@login_required
def enviar_usuarios_premium_correo(request):
    try:
        usuarios = obtener_usuarios_premium_activos()
        pdf_context = {**_contexto_base(request), 'usuarios': usuarios}
        headers = ['Nombre', 'Email', 'Plan', 'Registro']
        
        def write_excel_data(ws):
            for i, u in enumerate(usuarios, 2):
                ws.cell(row=i, column=1, value=u.get('nombre_completo', '—'))
                ws.cell(row=i, column=2, value=u.get('email', '—'))
                ws.cell(row=i, column=3, value=u.get('estadoPlan', '—'))
                ws.cell(row=i, column=4, value=str(u.get('fechaRegistro', '—')))
                
        return _dispatch_report_email(
            request=request,
            pdf_template='Suscripciones/reportes/pdf_usuarios_premium.html',
            pdf_context=pdf_context,
            pdf_filename='Usuarios_Premium_Activos.pdf',
            excel_headers=headers,
            write_excel_data_fn=write_excel_data,
            excel_filename='Usuarios_Premium_Activos.xlsx'
        )
    except Exception as e:
        return JsonResponse({'status': 'error', 'error': str(e)}, status=500)

@login_required
def exportar_usuarios_premium_excel(request):
    try:
        import openpyxl
        usuarios = obtener_usuarios_premium_activos()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Premium Activos'
        headers = ['Nombre', 'Email', 'Plan', 'Registro']
        for i, u in enumerate(usuarios, 2):
            ws.cell(row=i, column=1, value=u.get('nombre_completo', '—'))
            ws.cell(row=i, column=2, value=u.get('email', '—'))
            ws.cell(row=i, column=3, value=u.get('estadoPlan', '—'))
            ws.cell(row=i, column=4, value=str(u.get('fechaRegistro', '—')))
        _estilo_excel_sheet(ws, headers)
        return _build_excel_response(wb, 'Usuarios_Premium_Activos.xlsx')
    except Exception as e:
        return HttpResponse(f"Error: {str(e)}", status=500)


# ── REPORTE: CANDIDATOS A PLAN PREMIUM ──
def obtener_candidatos_premium():
    users = list(db["usuarios"].find({
        "estadoPlan": "Free",
        "estadoActivo": "Activo"
    }).sort("fechaRegistro", -1))
    for u in users:
        u["id"] = str(u.get("_id"))
        u["nombre_completo"] = f"{u.get('nombre', '')} {u.get('apellido', '')}"
        if isinstance(u.get("fechaRegistro"), str):
            u["fechaRegistro"] = parse_date(u["fechaRegistro"])
    return users

@login_required
def reporte_usuarios_free(request):
    try:
        usuarios = obtener_candidatos_premium()
        usuarios, total, page, total_pages, per_page = _paginate(request, usuarios)
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")
        usuarios = []
        total = page = total_pages = per_page = 0
    return render(request, 'Suscripciones/reportes/reporte_usuarios_free.html', {
        'usuarios': usuarios, 'page': page, 'total_pages': total_pages, 'total': total, 'per_page': per_page,
    })

@login_required
def exportar_usuarios_free_pdf(request):
    try:
        usuarios = obtener_candidatos_premium()
        context = {**_contexto_base(request), 'usuarios': usuarios}
        return _render_pdf(request, 'Suscripciones/reportes/pdf_usuarios_free.html', context, 'Candidatos_Plan_Premium.pdf')
    except Exception as e:
        return HttpResponse(f"Error: {str(e)}", status=500)

@login_required
def enviar_usuarios_free_correo(request):
    try:
        usuarios = obtener_candidatos_premium()
        pdf_context = {**_contexto_base(request), 'usuarios': usuarios}
        headers = ['Nombre', 'Email', 'Plan', 'Registro']
        
        def write_excel_data(ws):
            for i, u in enumerate(usuarios, 2):
                ws.cell(row=i, column=1, value=u.get('nombre_completo', '—'))
                ws.cell(row=i, column=2, value=u.get('email', '—'))
                ws.cell(row=i, column=3, value=u.get('estadoPlan', '—'))
                ws.cell(row=i, column=4, value=str(u.get('fechaRegistro', '—')))
                
        return _dispatch_report_email(
            request=request,
            pdf_template='Suscripciones/reportes/pdf_usuarios_free.html',
            pdf_context=pdf_context,
            pdf_filename='Candidatos_Plan_Premium.pdf',
            excel_headers=headers,
            write_excel_data_fn=write_excel_data,
            excel_filename='Candidatos_Plan_Premium.xlsx'
        )
    except Exception as e:
        return JsonResponse({'status': 'error', 'error': str(e)}, status=500)

@login_required
def exportar_usuarios_free_excel(request):
    try:
        import openpyxl
        usuarios = obtener_candidatos_premium()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Candidatos Premium'
        headers = ['Nombre', 'Email', 'Plan', 'Registro']
        for i, u in enumerate(usuarios, 2):
            ws.cell(row=i, column=1, value=u.get('nombre_completo', '—'))
            ws.cell(row=i, column=2, value=u.get('email', '—'))
            ws.cell(row=i, column=3, value=u.get('estadoPlan', '—'))
            ws.cell(row=i, column=4, value=str(u.get('fechaRegistro', '—')))
        _estilo_excel_sheet(ws, headers)
        return _build_excel_response(wb, 'Candidatos_Plan_Premium.xlsx')
    except Exception as e:
        return HttpResponse(f"Error: {str(e)}", status=500)


# ── REPORTE: INTENTOS DE ACCESO FALLIDOS ──
def obtener_accesos_fallidos():
    docs = list(db["auditoriaAcceso"].find({"resultado": "Fallido"}).sort("fechaHora", -1))
    user_ids = {d.get("idUsuario") for d in docs if d.get("idUsuario")}
    users = db["usuarios"].find({"id": {"$in": list(user_ids)}}, {"id": 1, "nombre": 1, "apellido": 1})
    user_names = {u["id"]: f"{u.get('nombre', '')} {u.get('apellido', '')}" for u in users}
    for d in docs:
        d["id"] = str(d.get("_id"))
        d["usuario_nombre"] = user_names.get(d.get("idUsuario"), f"Usuario {d.get('idUsuario', '—')}")
        if isinstance(d.get("fechaHora"), str):
            d["fechaHora"] = parse_date(d["fechaHora"])
    return docs

@login_required
def reporte_accesos_fallidos(request):
    try:
        accesos = obtener_accesos_fallidos()
        accesos, total, page, total_pages, per_page = _paginate(request, accesos)
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")
        accesos = []
        total = page = total_pages = per_page = 0
    return render(request, 'Suscripciones/reportes/reporte_accesos_fallidos.html', {
        'accesos': accesos, 'page': page, 'total_pages': total_pages, 'total': total, 'per_page': per_page,
    })

@login_required
def exportar_accesos_fallidos_pdf(request):
    try:
        accesos = obtener_accesos_fallidos()
        context = {**_contexto_base(request), 'accesos': accesos}
        return _render_pdf(request, 'Suscripciones/reportes/pdf_accesos_fallidos.html', context, 'Accesos_Fallidos_Audit.pdf')
    except Exception as e:
        return HttpResponse(f"Error: {str(e)}", status=500)

@login_required
def enviar_accesos_fallidos_correo(request):
    try:
        accesos = obtener_accesos_fallidos()
        pdf_context = {**_contexto_base(request), 'accesos': accesos}
        headers = ['Usuario', 'Acción', 'IP', 'Fecha']
        
        def write_excel_data(ws):
            for i, d in enumerate(accesos, 2):
                ws.cell(row=i, column=1, value=d.get('usuario_nombre', '—'))
                ws.cell(row=i, column=2, value=d.get('accion', '—'))
                ws.cell(row=i, column=3, value=d.get('ipOrigen', '—'))
                ws.cell(row=i, column=4, value=str(getattr(d, 'fechaHora', d.get('fechaHora', '—'))[:19] if hasattr(d, 'fechaHora') else '—'))
                
        return _dispatch_report_email(
            request=request,
            pdf_template='Suscripciones/reportes/pdf_accesos_fallidos.html',
            pdf_context=pdf_context,
            pdf_filename='Accesos_Fallidos_Audit.pdf',
            excel_headers=headers,
            write_excel_data_fn=write_excel_data,
            excel_filename='Accesos_Fallidos_Audit.xlsx'
        )
    except Exception as e:
        return JsonResponse({'status': 'error', 'error': str(e)}, status=500)

@login_required
def exportar_accesos_fallidos_excel(request):
    try:
        import openpyxl
        accesos = obtener_accesos_fallidos()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Accesos Fallidos'
        headers = ['Usuario', 'Acción', 'IP', 'Fecha']
        for i, d in enumerate(accesos, 2):
            ws.cell(row=i, column=1, value=d.get('usuario_nombre', '—'))
            ws.cell(row=i, column=2, value=d.get('accion', '—'))
            ws.cell(row=i, column=3, value=d.get('ipOrigen', '—'))
            ws.cell(row=i, column=4, value=str(getattr(d, 'fechaHora', d.get('fechaHora', '—'))[:19] if hasattr(d, 'fechaHora') else '—'))
        _estilo_excel_sheet(ws, headers)
        return _build_excel_response(wb, 'Accesos_Fallidos_Audit.xlsx')
    except Exception as e:
        return HttpResponse(f"Error: {str(e)}", status=500)


# ── REPORTE: ACCIONES DE ADMINISTRACIÓN ──
def obtener_acciones_admin():
    docs = list(db["auditoriaAcceso"].find({
        "accion": {
            "$in": [
                "CAMBIO_CONTRASENA",
                "CAMBIO_PLAN",
                "ADMIN_ACCION"
            ]
        }
    }).sort("fechaHora", -1))
    user_ids = {d.get("idUsuario") for d in docs if d.get("idUsuario")}
    users = db["usuarios"].find({"id": {"$in": list(user_ids)}}, {"id": 1, "nombre": 1, "apellido": 1})
    user_names = {u["id"]: f"{u.get('nombre', '')} {u.get('apellido', '')}" for u in users}
    for d in docs:
        d["id"] = str(d.get("_id"))
        d["usuario_nombre"] = user_names.get(d.get("idUsuario"), f"Usuario {d.get('idUsuario', '—')}")
        if isinstance(d.get("fechaHora"), str):
            d["fechaHora"] = parse_date(d["fechaHora"])
    return docs

@login_required
def reporte_acciones_admin(request):
    try:
        acciones = obtener_acciones_admin()
        acciones, total, page, total_pages, per_page = _paginate(request, acciones)
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")
        acciones = []
        total = page = total_pages = per_page = 0
    return render(request, 'Suscripciones/reportes/reporte_acciones_admin.html', {
        'acciones': acciones, 'page': page, 'total_pages': total_pages, 'total': total, 'per_page': per_page,
    })

@login_required
def exportar_acciones_admin_pdf(request):
    try:
        acciones = obtener_acciones_admin()
        context = {**_contexto_base(request), 'acciones': acciones}
        return _render_pdf(request, 'Suscripciones/reportes/pdf_acciones_admin.html', context, 'Acciones_Administracion_Audit.pdf')
    except Exception as e:
        return HttpResponse(f"Error: {str(e)}", status=500)

@login_required
def enviar_acciones_admin_correo(request):
    try:
        acciones = obtener_acciones_admin()
        pdf_context = {**_contexto_base(request), 'acciones': acciones}
        headers = ['Usuario', 'Acción', 'IP', 'Fecha']
        
        def write_excel_data(ws):
            for i, d in enumerate(acciones, 2):
                ws.cell(row=i, column=1, value=d.get('usuario_nombre', '—'))
                ws.cell(row=i, column=2, value=d.get('accion', '—'))
                ws.cell(row=i, column=3, value=d.get('ipOrigen', '—'))
                ws.cell(row=i, column=4, value=str(d.get('fechaHora', '—'))[:19])
                
        return _dispatch_report_email(
            request=request,
            pdf_template='Suscripciones/reportes/pdf_acciones_admin.html',
            pdf_context=pdf_context,
            pdf_filename='Acciones_Administracion_Audit.pdf',
            excel_headers=headers,
            write_excel_data_fn=write_excel_data,
            excel_filename='Acciones_Administracion_Audit.xlsx'
        )
    except Exception as e:
        return JsonResponse({'status': 'error', 'error': str(e)}, status=500)

@login_required
def exportar_acciones_admin_excel(request):
    try:
        import openpyxl
        acciones = obtener_acciones_admin()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Acciones Admin'
        headers = ['Usuario', 'Acción', 'IP', 'Fecha']
        for i, d in enumerate(acciones, 2):
            ws.cell(row=i, column=1, value=d.get('usuario_nombre', '—'))
            ws.cell(row=i, column=2, value=d.get('accion', '—'))
            ws.cell(row=i, column=3, value=d.get('ipOrigen', '—'))
            ws.cell(row=i, column=4, value=str(d.get('fechaHora', '—'))[:19])
        _estilo_excel_sheet(ws, headers)
        return _build_excel_response(wb, 'Acciones_Administracion_Audit.xlsx')
    except Exception as e:
        return HttpResponse(f"Error: {str(e)}", status=500)


# ══════════════════════════════════════════
#  DASHBOARD DEL ANALISTA
# ══════════════════════════════════════════

def _dashboard_kpis():
    """Calcula todos los KPIs y datos de gráficas para el dashboard. Retorna un dict."""
    from catalogo.views import obtener_ranking_popularidad_mensual
    data = {}

    # ── KPI 1: Canción #1 del mes ──
    try:
        ranking = obtener_ranking_popularidad_mensual()
        if ranking:
            data['cancion_top'] = ranking[0].get('tituloCancion') or ranking[0].get('nombre') or '—'
            data['cancion_top_artista'] = ranking[0].get('nombreArtistico', '—')
        else:
            primera = db['Cancion'].find_one({}, {'tituloCancion': 1})
            data['cancion_top'] = primera.get('tituloCancion', '—') if primera else '—'
            data['cancion_top_artista'] = '—'
    except Exception:
        data['cancion_top'] = '—'
        data['cancion_top_artista'] = '—'

    # ── KPI 2: Total usuarios premium activos ──
    try:
        data['total_premium'] = usuarios_col.count_documents({'estadoPlan': 'Premium', 'estadoActivo': 'Activo'})
    except Exception:
        data['total_premium'] = 0

    # ── KPI 3: Suscripciones que vencen en 5 días ──
    try:
        hoy = datetime.now()
        en_5_dias = hoy + timedelta(days=5)
        docs_sus = list(suscripciones_col.find({'estado': {'$in': ['Activa', 'activa']}}, {'fechaFin': 1}))
        count = 0
        for s in docs_sus:
            fecha_fin = s.get('fechaFin')
            if isinstance(fecha_fin, str):
                try:
                    fecha_fin = datetime.fromisoformat(fecha_fin)
                except Exception:
                    continue
            if fecha_fin and hoy <= fecha_fin <= en_5_dias:
                count += 1
        data['vencen_pronto'] = count
    except Exception:
        data['vencen_pronto'] = 0

    # ── KPI 4: Total géneros ──
    try:
        data['total_generos'] = db['Genero'].count_documents({})
    except Exception:
        data['total_generos'] = 0

    # ── KPI 5: Ingresos estimados del mes ──
    try:
        sus_activas = list(suscripciones_col.find({'estado': {'$in': ['Activa', 'activa']}}, {'tipoSuscripcion.precio': 1}))
        total_ingresos = 0
        for s in sus_activas:
            tipo = s.get('tipoSuscripcion', {})
            if isinstance(tipo, dict):
                total_ingresos += tipo.get('precio', 0) or 0
        data['ingresos_estimados'] = round(total_ingresos, 2)
    except Exception:
        data['ingresos_estimados'] = 0

    # ── KPI 6: % conversión Free → Premium ──
    try:
        total_users = usuarios_col.count_documents({})
        if total_users > 0:
            data['conversion_pct'] = round((data['total_premium'] / total_users) * 100, 1)
        else:
            data['conversion_pct'] = 0
    except Exception:
        data['conversion_pct'] = 0

    # ── Datos para gráfico de pastel: distribución por plan ──
    try:
        pipeline = [
            {"$group": {"_id": "$estadoPlan", "count": {"$sum": 1}}}
        ]
        plan_dist = list(usuarios_col.aggregate(pipeline))
        data['plan_labels'] = json.dumps([d['_id'] or 'Sin plan' for d in plan_dist])
        data['plan_data'] = json.dumps([d['count'] for d in plan_dist])
    except Exception:
        data['plan_labels'] = '[]'
        data['plan_data'] = '[]'

    # ── Datos para gráfico de barras: top 5 canciones ──
    try:
        ranking = obtener_ranking_popularidad_mensual()
        top5 = ranking[:5] if ranking else []
        data['top5_labels'] = json.dumps([t.get('tituloCancion', '—')[:20] for t in top5])
        data['top5_data'] = json.dumps([t.get('total_escuchas_mes', 0) for t in top5])
    except Exception:
        data['top5_labels'] = '[]'
        data['top5_data'] = '[]'

    return data


@login_required
def analista_dashboard(request):
    context = _dashboard_kpis()
    return render(request, 'Suscripciones/dashboard_analista.html', context)


@login_required
def analista_dashboard_data(request):
    """Endpoint JSON para refrescar KPIs vía AJAX."""
    data = _dashboard_kpis()
    return JsonResponse(data)


# ── VISTAS Y APIS DE REPORTES PERSONALIZADOS ──

def _es_query_segura(pipeline):
    """
    Verifica que el pipeline de agregación sea seguro (solo lectura).
    Evita etapas destructivas de MongoDB.
    """
    etapas_prohibidas = {'$out', '$merge', '$writeConcern', '$collStats'}
    for etapa in pipeline:
        if not isinstance(etapa, dict):
            return False
        for key in etapa.keys():
            if key in etapas_prohibidas:
                return False
    return True


@login_required
def crear_reporte_hub(request):
    return render(request, 'Suscripciones/reportes/crear_reporte_hub.html')


@login_required
def crear_reporte_mongodb(request):
    return render(request, 'Suscripciones/reportes/crear_reporte_mongodb.html')


@login_required
def crear_reporte_powerbi(request):
    if request.method == 'POST':
        import json
        data = json.loads(request.body)
        workspace_id = data.get('workspace_id')
        report_id = data.get('report_id')
        
        # En producción real se negociaría el token seguro con MSAL/Azure AD.
        # Retornamos un token dummy de demostración.
        return JsonResponse({
            'status': 'success',
            'embed_token': 'dummy-embed-token-powerbi-1234567890abcdef',
            'embed_url': f'https://app.powerbi.com/reportEmbed?reportId={report_id}&groupId={workspace_id}',
        })
    return render(request, 'Suscripciones/reportes/crear_reporte_powerbi.html')


@login_required
def crear_reporte_ia(request):
    return render(request, 'Suscripciones/reportes/crear_reporte_ia.html')


@login_required
def ver_reporte_personalizado(request, report_id):
    import json
    from bson import ObjectId, json_util
    try:
        report_doc = db['reportes_personalizados'].find_one({'_id': ObjectId(report_id)})
        if not report_doc:
            return redirect('crear_reporte_hub')
            
        metodo = report_doc.get('metodo')
        config = report_doc.get('configuracion', {})
        
        context = {
            'report_id': report_id,
            'report_title': report_doc.get('titulo'),
            'report_config': json_util.dumps(config)
        }
        
        if metodo == 'mongodb':
            return render(request, 'Suscripciones/reportes/crear_reporte_mongodb.html', context)
        elif metodo == 'powerbi':
            return render(request, 'Suscripciones/reportes/crear_reporte_powerbi.html', context)
        elif metodo == 'ia':
            # Ejecutamos el pipeline guardado para enviar los resultados pre-renderizados
            coleccion = config.get('coleccion')
            pipeline = config.get('pipeline', [])
            resultados = list(db[coleccion].aggregate(pipeline))
            context['report_results'] = json_util.dumps(resultados)
            return render(request, 'Suscripciones/reportes/crear_reporte_ia.html', context)
            
    except Exception as e:
        print("Error al cargar reporte personalizado:", e)
    return redirect('crear_reporte_hub')


@login_required
def ejecutar_query_mongo_api(request):
    if request.method == 'POST':
        import json
        from bson import json_util
        try:
            data = json.loads(request.body)
            coleccion = data.get('coleccion')
            pipeline = json.loads(data.get('pipeline'))
            
            # Forzar limit para prevenir sobrecarga de memoria
            tiene_limit = False
            for etapa in pipeline:
                if '$limit' in etapa:
                    tiene_limit = True
                    if etapa['$limit'] > 1000:
                        etapa['$limit'] = 1000
            
            if not tiene_limit:
                pipeline.append({'$limit': 200})
                
            if not _es_query_segura(pipeline):
                return JsonResponse({'error': 'La consulta contiene etapas no permitidas.'})
                
            resultados = list(db[coleccion].aggregate(pipeline))
            return HttpResponse(json_util.dumps(resultados), content_type='application/json')
        except Exception as e:
            return JsonResponse({'error': str(e)})
    return JsonResponse({'error': 'Método no permitido'})


@login_required
def ia_generar_query_api(request):
    if request.method == 'POST':
        import os
        import json
        import requests
        from bson import json_util
        try:
            data = json.loads(request.body)
            prompt_usuario = data.get('prompt')
            
            gemini_key = getattr(settings, 'GEMINI_API_KEY', os.environ.get('GEMINI_API_KEY', ''))
            
            prompt_sistema = (
                "Actúa como un traductor experto de lenguaje natural a pipelines de agregación de MongoDB para el proyecto Cénit.\n"
                "Nuestras colecciones y sus estructuras reales en la base de datos son:\n"
                "- 'Cancion': { cancion_id: int, tituloCancion: str, album_id: int, genero_id: int, escuchas: int } (Nota: genero_id relaciona con Genero.genero_id; album_id relaciona con Album.album_id)\n"
                "- 'Artista': { artista_id: int, nombreArtistico: str, paisOrigen: str, biografia: str, estadoActivo: str }\n"
                "- 'Album': { album_id: int, tituloAlbum: str, artista_id: int, fechaLanzamiento: date }\n"
                "- 'Genero': { genero_id: int, nombreGenero: str }\n"
                "- 'usuarios': { id: int, nombre: str, apellido: str, email: str, rol: { nombreRol: str } }\n"
                "- 'suscripciones': { idSuscripcion: int, idUsuario: int, plan: str, estado: str }\n"
                "- 'tipoSuscripciones': { tipo_id: int, nombrePlan: str, precio: float, moneda: str, duracion: int } (Nota: relaciona tipoSuscripciones.tipo_id con suscripciones.plan)\n"
                "- 'promociones': { idPromocion: int, codigo: str, descuento: float, estado: str }\n"
                "- 'playlists': { idPlaylist: int, idUsuario: int, nombre: str, canciones: array }\n"
                "- 'notificaciones': { tipoNotif: str, mensaje: str, fechaEnvio: str, idUsuario: int, idPromocion: int }\n"
                "- 'auditoriaAcceso': { idAuditoria: int, email: str, resultado: str, fechaHora: date }\n"
                "- 'estadisticasDiarias': { fecha: date, escuchasTotales: int, usuariosActivos: int }\n"
                "- 'seguimientos': { idUsuario: int, idArtista: int, nombreArtista: str, fechaSeguimiento: str, activo: int } (relaciona qué usuarios siguen a qué artistas)\n"
                "- 'cancionesFavoritas': { idUsuario: int, idCancion: int, tituloCancion: str, artista: str, album: str, fechaLike: str } (relaciona canciones con me gusta de cada usuario)\n"
                "- 'roles': { _id: int, nombreRol: str, descripcion: str, activo: int, fechaCreacion: str }\n\n"
                "REGLA CRÍTICA DE BÚSQUEDA DE TEXTO:\n"
                "Como las consultas en MongoDB son estrictamente sensibles a mayúsculas y minúsculas (case-sensitive), cuando filtres por campos de texto (ej. nombres de artistas, títulos de canciones, planes, etc.), debes usar expresiones regulares insensibles a mayúsculas y minúsculas con $regex y $options: 'i'. Por ejemplo:\n"
                "{ \"artistaInfo.nombreArtistico\": { \"$regex\": \"^Lana del Rey$\", \"$options\": \"i\" } } en lugar de una igualdad directa.\n\n"
                "Retorna ÚNICAMENTE un objeto JSON válido con el siguiente formato, sin texto descriptivo ni formato markdown:\n"
                "{\n"
                "  \"titulo\": \"Título representativo para el reporte\",\n"
                "  \"coleccion\": \"Nombre exacto de la colección de MongoDB a consultar\",\n"
                "  \"pipeline\": [ ...lista de fases de agregación... ]\n"
                "}\n"
                "Asegúrate de que la colección empiece por mayúscula si es Cancion, Artista, Album o Genero, y minúscula para las demás."
            )
            
            generated_json = None
            if gemini_key:
                url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={gemini_key}"
                payload = {
                    "contents": [{
                        "parts": [
                            {"text": prompt_sistema},
                            {"text": f"Pregunta del Analista: {prompt_usuario}"}
                        ]
                    }],
                    "generationConfig": {
                        "responseMimeType": "application/json"
                    }
                }
                response = requests.post(url, json=payload, headers={"Content-Type": "application/json"})
                if response.status_code == 200:
                    res_json = response.json()
                    try:
                        text = res_json['candidates'][0]['content']['parts'][0]['text']
                        generated_json = json.loads(text)
                    except Exception as e:
                        print("Error al analizar respuesta de Gemini:", e)
            
            # Mock Fallback local
            if not generated_json:
                p_lower = prompt_usuario.lower()
                if 'rock' in p_lower or 'cancion' in p_lower:
                    generated_json = {
                        "titulo": "Canciones de Rock Populares (IA Mock)",
                        "coleccion": "Cancion",
                        "pipeline": [
                            { "$match": { "genero": "Rock" } },
                            { "$sort": { "escuchas": -1 } },
                            { "$limit": 10 }
                        ]
                    }
                elif 'usuario' in p_lower or 'premium' in p_lower:
                    generated_json = {
                        "titulo": "Usuarios Registrados Recientes (IA Mock)",
                        "coleccion": "usuarios",
                        "pipeline": [
                            { "$sort": { "id": -1 } },
                            { "$limit": 10 }
                        ]
                    }
                else:
                    generated_json = {
                        "titulo": "Canciones Generales del Catálogo (IA Mock)",
                        "coleccion": "Cancion",
                        "pipeline": [
                            { "$sort": { "escuchas": -1 } },
                            { "$limit": 10 }
                        ]
                    }
            
            coleccion = generated_json.get('coleccion')
            pipeline = generated_json.get('pipeline', [])
            
            # Limit de seguridad
            pipeline.append({'$limit': 200})
            if not _es_query_segura(pipeline):
                return JsonResponse({'status': 'error', 'message': 'Pipeline generado no seguro.'})
                
            resultados = list(db[coleccion].aggregate(pipeline))
            
            return JsonResponse({
                'status': 'success',
                'titulo': generated_json.get('titulo'),
                'coleccion': coleccion,
                'pipeline': pipeline[:-1],  # Remove limit para mostrar limpio
                'pregunta': prompt_usuario,
                'resultados': json.loads(json_util.dumps(resultados))
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'})


@login_required
def guardar_reporte_api(request):
    if request.method == 'POST':
        import json
        import datetime
        try:
            data = json.loads(request.body)
            titulo = data.get('titulo')
            metodo = data.get('metodo')
            configuracion = data.get('configuracion')
            
            doc = {
                "titulo": titulo,
                "creador": request.user.username,
                "metodo": metodo,
                "configuracion": configuracion,
                "fechaCreacion": datetime.datetime.now()
            }
            
            db['reportes_personalizados'].insert_one(doc)
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'error': str(e)})
    return JsonResponse({'status': 'error', 'error': 'Método no permitido'})


@login_required
def exportar_reporte_dinamico_pdf(request):
    if request.method == 'POST':
        import json
        import datetime
        from bson import ObjectId
        try:
            coleccion = request.POST.get('coleccion')
            pipeline_str = request.POST.get('pipeline')
            titulo = request.POST.get('titulo', 'Reporte Personalizado')
            
            pipeline = json.loads(pipeline_str)
            resultados = list(db[coleccion].aggregate(pipeline))
            
            # Obtener columnas únicas
            columnas = []
            for item in resultados:
                for key in item.keys():
                    if key not in columnas:
                        columnas.append(key)
            
            # Formatear filas como lista de listas
            datos_lista = []
            for item in resultados:
                fila = []
                for col in columnas:
                    val = item.get(col, '-')
                    if isinstance(val, ObjectId):
                        val = str(val)
                    elif isinstance(val, (datetime.datetime, datetime.date)):
                        val = val.strftime('%d/%m/%Y %H:%M') if isinstance(val, datetime.datetime) else val.strftime('%d/%m/%Y')
                    elif isinstance(val, (dict, list)):
                        val = json.dumps(val, default=str)
                    fila.append(val)
                datos_lista.append(fila)
            
            context = {
                'titulo': titulo,
                'coleccion': coleccion,
                'fecha': datetime.datetime.now().strftime('%d/%m/%Y %H:%M'),
                'columnas': columnas,
                'datos_lista': datos_lista,
                'usuario_generador': request.user.username
            }
            
            from weasyprint import HTML
            html_string = render_to_string('Suscripciones/reportes/pdf_reporte_dinamico.html', context)
            pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()
            
            response = HttpResponse(pdf_file, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="reporte_dinamico.pdf"'
            return response
        except Exception as e:
            return HttpResponse(f"Error al generar PDF: {str(e)}", status=500)
    return HttpResponse("Método no permitido", status=405)


@login_required
def exportar_reporte_dinamico_excel(request):
    if request.method == 'POST':
        import json
        import openpyxl
        import datetime
        from bson import ObjectId
        try:
            coleccion = request.POST.get('coleccion')
            pipeline_str = request.POST.get('pipeline')
            titulo = request.POST.get('titulo', 'Reporte_Personalizado')
            pipeline = json.loads(pipeline_str)
            resultados = list(db[coleccion].aggregate(pipeline))
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Resultados'
            
            if not resultados:
                return _build_excel_response(wb, 'Reporte_Vacio.xlsx')
                
            columnas = []
            for item in resultados:
                for key in item.keys():
                    if key not in columnas:
                        columnas.append(key)
            
            # Escribir filas de datos
            for row_idx, item in enumerate(resultados, 2):
                for col_idx, col_name in enumerate(columnas, 1):
                    val = item.get(col_name, '')
                    if isinstance(val, ObjectId):
                        val = str(val)
                    elif isinstance(val, (datetime.datetime, datetime.date)):
                        val = val.strftime('%d/%m/%Y %H:%M') if isinstance(val, datetime.datetime) else val.strftime('%d/%m/%Y')
                    elif isinstance(val, (dict, list)):
                        val = json.dumps(val, default=str)
                    ws.cell(row=row_idx, column=col_idx, value=val)
                    
            # Aplicar formato de hoja premium
            _estilo_excel_sheet(ws, columnas)
            
            # Sanitizar nombre del archivo
            filename = f"{titulo.replace(' ', '_')}.xlsx"
            return _build_excel_response(wb, filename)
        except Exception as e:
            return HttpResponse(f"Error al generar Excel: {str(e)}", status=500)
    return HttpResponse("Método no permitido", status=405)


@login_required
def eliminar_reporte(request, report_id):
    from bson import ObjectId
    try:
        result = db['reportes_personalizados'].delete_one({'_id': ObjectId(report_id)})
        if result.deleted_count > 0:
            messages.success(request, "El reporte personalizado fue eliminado exitosamente.")
        else:
            messages.error(request, "No se encontró el reporte a eliminar.")
    except Exception as e:
        messages.error(request, f"Error al eliminar el reporte: {str(e)}")
    return redirect('analista_dashboard')


@login_required
def enviar_reporte_dinamico_correo(request):
    import json
    import datetime
    from bson import ObjectId
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'error': 'Método no permitido'}, status=405)
        
    try:
        data = json.loads(request.body)
        coleccion = data.get('coleccion')
        pipeline_str = data.get('pipeline')
        titulo = data.get('titulo', 'Reporte_Personalizado')
        
        # Parse pipeline y ejecutar agregación
        pipeline = json.loads(pipeline_str)
        resultados = list(db[coleccion].aggregate(pipeline))
        
        # 1. Configurar columnas y contexto para PDF
        columnas = []
        if resultados:
            for item in resultados:
                for key in item.keys():
                    if key not in columnas:
                        columnas.append(key)
                        
        pdf_context = {
            **_contexto_base(request),
            'titulo': titulo,
            'coleccion': coleccion,
            'columnas': columnas,
            'resultados': resultados,
            'fecha': datetime.datetime.now()
        }
        
        # 2. Configurar callback para Excel
        def write_excel_data(ws):
            for row_idx, item in enumerate(resultados, 2):
                for col_idx, col_name in enumerate(columnas, 1):
                    val = item.get(col_name, '')
                    if isinstance(val, ObjectId):
                        val = str(val)
                    elif isinstance(val, (datetime.datetime, datetime.date)):
                        val = val.strftime('%d/%m/%Y %H:%M') if isinstance(val, datetime.datetime) else val.strftime('%d/%m/%Y')
                    elif isinstance(val, (dict, list)):
                        val = json.dumps(val, default=str)
                    ws.cell(row=row_idx, column=col_idx, value=val)
                    
        return _dispatch_report_email(
            request=request,
            pdf_template='Suscripciones/reportes/pdf_reporte_dinamico.html',
            pdf_context=pdf_context,
            pdf_filename=f"{titulo.replace(' ', '_')}.pdf",
            excel_headers=columnas,
            write_excel_data_fn=write_excel_data,
            excel_filename=f"{titulo.replace(' ', '_')}.xlsx"
        )
    except Exception as e:
        return JsonResponse({'status': 'error', 'error': f"Error al procesar envío de correo: {str(e)}"}, status=500)